"""
فحص ذاتي بلا اتصال بالنظام: يتأكد أن الاستخراج والتصنيف وكتابة الإكسل
وقراءة الـPDF تعمل على هذا الجهاز قبل تشغيل الأتمتة الحقيقية.
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from .classify import GASTRO, PHYSIO, UNKNOWN, category_label, classify
from .config import load_selectors
from .excel import COLUMNS, STATUS_COLUMN, append_rows, existing_task_numbers, update_statuses
from .parse import (
    merge_sources,
    normalize_date,
    normalize_digits,
    parse_arabic_date,
    parse_pdf_fields,
    parse_sender,
    parse_task_number,
    parse_task_title,
)
from .pdf import extract_pdf_text

results: list[tuple[str, bool, str]] = []


def check(name: str, fn) -> None:
    try:
        fn()
        results.append((name, True, ""))
    except Exception as error:  # noqa: BLE001
        results.append((name, False, str(error)))


def build_sample_pdf(target: Path, line: str) -> Path:
    """يبني ملف PDF صغير صالح بطبقة نصية، لاختبار مسار القراءة فعليا."""
    escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    content = f"BT /F1 12 Tf 60 720 Td ({escaped}) Tj ET"

    objects = [
        "<</Type/Catalog/Pages 2 0 R>>",
        "<</Type/Pages/Kids[3 0 R]/Count 1>>",
        "<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
        "/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>",
        "<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
        f"<</Length {len(content)}>>\nstream\n{content}\nendstream",
    ]

    pdf = "%PDF-1.4\n"
    offsets = []
    for index, body in enumerate(objects, start=1):
        offsets.append(len(pdf.encode("latin-1")))
        pdf += f"{index} 0 obj\n{body}\nendobj\n"

    xref_offset = len(pdf.encode("latin-1"))
    pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for offset in offsets:
        pdf += f"{offset:010d} 00000 n \n"
    pdf += f"trailer\n<</Size {len(objects) + 1}/Root 1 0 R>>\nstartxref\n{xref_offset}\n%%EOF"

    target.write_bytes(pdf.encode("latin-1"))
    return target


SAMPLE = """مستشفى الملك فهد
رقم المعاملة : 44120
اسم المريض : أحمد محمد العتيبي
رقم الملف : ٧٨٩٤٥٦
القسم : العلاج الطبيعي
الطبيب المحول : د. سارة القحطاني
التشخيص : آلام أسفل الظهر مع ضعف في عضلات الطرف السفلي
نوع الطلب : تحويل لجلسات علاج طبيعي
التاريخ : 2026-07-27"""


def main() -> int:
    temp = Path(tempfile.mkdtemp(prefix="daily-tasks-selftest-"))
    selectors = load_selectors()

    # عناوين حقيقية منقولة حرفيا من صفحة المهام الواردة في ttn.ksu.edu.sa
    def title_five_parts():
        parsed = parse_task_title(
            "طلب تقرير انجليزي - خلود سالم محمد القحطاني - 1065170 - ماجد الماضي - الجهاز الهضمي"
        )
        assert parsed["requestType"] == "طلب تقرير انجليزي", parsed
        assert parsed["patientName"] == "خلود سالم محمد القحطاني", parsed
        assert parsed["fileNumber"] == "1065170", parsed
        assert parsed["doctor"] == "ماجد الماضي", parsed
        assert parsed["department"] == "الجهاز الهضمي", parsed

    def title_four_parts():
        parsed = parse_task_title("طلب تقرير انجليزي - هويده جميل سنوسي صابر - 10389908 - العلاج الطبيعي")
        assert parsed["patientName"] == "هويده جميل سنوسي صابر", parsed
        assert parsed["fileNumber"] == "10389908", parsed
        assert parsed["doctor"] == "", f"توقعنا بلا طبيب لكن ظهر: {parsed['doctor']}"
        assert parsed["department"] == "العلاج الطبيعي", parsed

    def classify_from_title():
        cases = [
            ("طلب تقرير انجليزي - دليل محسن هميجان المطيري - 10285310 - سعد الخويطر - الجهاز الهضمي", GASTRO),
            ("طلب تقرير انجليزي - حصه علي ابراهيم البيطار - 123727 - العلاج الطبيعي", PHYSIO),
            ("طلب تقرير انجليزي - نوال فرحان هريسان الرشيدي - 1201518 - نهله عزام - الجهاز الهضمي", GASTRO),
            ("طلب تقرير انجليزي - مسفر محمد عون الجبيري - 10392393 - مهند الطيب - الجهاز الهضمي", GASTRO),
            ("طلب تقرير انجليزي - حميد بخيت القثامي - 10567029 - الجهاز الهضمي", GASTRO),
        ]
        for title, expected in cases:
            verdict = classify(parse_task_title(title), title, selectors["classification"])
            assert verdict["category"] == expected, f"{title} صنف كـ{verdict['category']}"

    def sender_from_card():
        assert parse_sender("من: Munirah Aldosari إلى: أحمد ناصر فهد الأحمد") == "Munirah Aldosari"

    def task_number_from_card():
        assert parse_task_number("الرقم: 144800006202") == "144800006202"
        assert parse_task_number("الرقم: ١٤٤٨٠٠٠٠٦٩٠٠") == "144800006900"

    def hijri_date():
        assert parse_arabic_date("05 صفر 1448 الموافق 19 يوليو 2026") == "2026-07-19"
        assert parse_arabic_date("08 صفر 1448 الموافق 22 يوليو 2026") == "2026-07-22"
        assert parse_arabic_date("07 صفر 1448 الموافق 21 يوليو 2026") == "2026-07-21"

    def arabic_pdf_fields():
        fields = parse_pdf_fields(SAMPLE)
        assert fields["patientName"] == "أحمد محمد العتيبي", fields
        assert fields["fileNumber"] == "789456", f"رقم الملف المستخرج: {fields['fileNumber']}"
        assert fields["taskNumber"] == "44120", fields
        assert fields["department"] == "العلاج الطبيعي", fields
        assert fields["documentDate"] == "2026-07-27", fields

    def one_line_pdf():
        one_line = (
            "Task No: 44120 Patient Name: Ahmed Al Otaibi File No: 789456 "
            "Department: Physiotherapy Diagnosis: Lower back pain Date: 2026-07-27"
        )
        fields = parse_pdf_fields(one_line)
        assert fields["patientName"] == "Ahmed Al Otaibi", fields["patientName"]
        assert fields["fileNumber"] == "789456", fields["fileNumber"]
        assert fields["taskNumber"] == "44120", fields["taskNumber"]
        assert fields["department"] == "Physiotherapy", fields["department"]

    def digits_and_dates():
        assert normalize_digits("١٢٣٤") == "1234"
        assert normalize_date("27/07/2026") == "2026-07-27"
        assert normalize_date("2026/7/5") == "2026-07-05"

    def classify_physio():
        verdict = classify(parse_pdf_fields(SAMPLE), SAMPLE, selectors["classification"])
        assert verdict["category"] == PHYSIO, verdict

    def classify_gastro():
        text = "القسم : الجهاز الهضمي\nنوع الطلب : منظار قولون تشخيصي"
        verdict = classify(parse_pdf_fields(text), text, selectors["classification"])
        assert verdict["category"] == GASTRO, verdict

    def classify_unknown():
        text = "طلب صرف مستلزمات مكتبية"
        verdict = classify(parse_pdf_fields(text), text, selectors["classification"])
        assert verdict["category"] == UNKNOWN, verdict
        assert category_label(verdict["category"]) == "غير مصنف"

    def source_priority():
        merged = merge_sources(
            row={"taskNumber": "9001", "patientName": ""},
            pdf={"patientName": "نورة السالم", "fileNumber": ""},
            ai={"fileNumber": "112233", "summary": "تحويل لجلسة تأهيل"},
        )
        assert merged["taskNumber"] == "9001", merged
        assert merged["patientName"] == "نورة السالم", merged
        assert merged["fileNumber"] == "112233", merged
        assert merged["summary"] == "تحويل لجلسة تأهيل", merged

    def read_real_pdf():
        path = build_sample_pdf(temp / "sample.pdf", "File No: 445566 Patient: Ahmed")
        text = extract_pdf_text(path)
        assert "445566" in text, f"النص المستخرج: {text}"

    def excel_roundtrip():
        path = temp / "physio.xlsx"
        row = {
            "taskNumber": "144800006202",
            "patientName": "خلود سالم محمد القحطاني",
            "fileNumber": "1065170",
            "department": "الجهاز الهضمي",
            "status": "مسجلة - بانتظار التنفيذ",
        }

        append_rows(path, "2026-07-27", [row])
        append_rows(path, "2026-07-28", [{**row, "taskNumber": "144800006900"}])

        seen = existing_task_numbers(path)
        assert {"144800006202", "144800006900"} <= seen, seen

        updated = update_statuses(path, "2026-07-27", {"144800006202": "تحت التنفيذ"})
        assert updated == 1, f"حدث {updated} صفا"

        workbook = load_workbook(path)
        assert workbook.sheetnames == ["2026-07-27", "2026-07-28"], workbook.sheetnames

        sheet = workbook["2026-07-27"]
        assert sheet.cell(row=1, column=1).value == COLUMNS[0][0]
        assert sheet.sheet_view.rightToLeft is True, "الشريحة ليست بترتيب من اليمين لليسار"
        assert sheet.cell(row=2, column=STATUS_COLUMN).value == "تحت التنفيذ"
        workbook.close()

    for name, fn in [
        ("قراءة عنوان المهمة بخمسة أجزاء (مع اسم الطبيب)", title_five_parts),
        ("قراءة عنوان المهمة بأربعة أجزاء (بلا اسم طبيب)", title_four_parts),
        ("التصنيف من عنوان المهمة مباشرة", classify_from_title),
        ("قراءة الجهة المرسلة من البطاقة", sender_from_card),
        ("قراءة رقم المهمة من البطاقة", task_number_from_card),
        ("قراءة التاريخ الميلادي بعد «الموافق»", hijri_date),
        ("استخراج الحقول من نص عربي", arabic_pdf_fields),
        ("PDF مستخرج كسطر واحد لا تبتلع فيه القيمة بقية الحقول", one_line_pdf),
        ("تحويل الأرقام العربية والتواريخ", digits_and_dates),
        ("تصنيف مريض علاج طبيعي", classify_physio),
        ("تصنيف مريض جهاز هضمي", classify_gastro),
        ("المعاملة الغامضة تذهب لملف غير مصنف", classify_unknown),
        ("أولوية المصادر: العنوان ثم PDF ثم الذكاء الاصطناعي", source_priority),
        ("قراءة نص من ملف PDF حقيقي", read_real_pdf),
        ("كتابة الإكسل بشريحة يومية وعدم تكرار المعاملات", excel_roundtrip),
    ]:
        check(name, fn)

    failed = [item for item in results if not item[1]]
    for name, ok, error in results:
        print(f"{'✔' if ok else '✘'}  {name}" + ("" if ok else f"\n    {error}"))

    print(f"\n{len(results) - len(failed)}/{len(results)} فحص ناجح.")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
