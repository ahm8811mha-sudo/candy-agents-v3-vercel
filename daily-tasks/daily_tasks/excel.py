"""كتابة ملفات الإكسل: شريحة مستقلة لكل يوم، ومنع تكرار المعاملات."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("م", "serial", 6),
    ("رقم المهمة", "taskNumber", 16),
    ("اسم المريض", "patientName", 30),
    ("رقم الملف", "fileNumber", 16),
    ("القسم", "department", 18),
    ("نوع الطلب / الموضوع", "requestType", 26),
    ("التشخيص", "diagnosis", 26),
    ("الطبيب المحول", "doctor", 22),
    ("الجهة المرسلة", "sender", 22),
    ("تاريخ المعاملة", "documentDate", 14),
    ("الملخص", "summary", 44),
    ("حالة المعاملة", "status", 16),
    ("مصدر البيانات", "source", 18),
    ("وقت التسجيل", "recordedAt", 12),
    ("رابط المهمة", "taskUrl", 40),
    ("ملف الـPDF", "pdfPath", 28),
]

KEYS = [key for _, key, _ in COLUMNS]
TASK_COLUMN = KEYS.index("taskNumber") + 1
STATUS_COLUMN = KEYS.index("status") + 1

HEADER_FILL = PatternFill("solid", fgColor="FF1F3B4D")
NEEDS_REVIEW_FILL = PatternFill("solid", fgColor="FFFFF3CD")


def _open_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def _get_day_sheet(workbook: Workbook, day_key: str):
    """شريحة مستقلة لكل يوم، فتبقى مهام الأربعاء منفصلة تماما عن مهام الثلاثاء."""
    if day_key in workbook.sheetnames:
        return workbook[day_key]

    sheet = workbook.create_sheet(day_key)
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"

    sheet.append([header for header, _, _ in COLUMNS])

    header_row = sheet[1]
    sheet.row_dimensions[1].height = 26
    for index, cell in enumerate(header_row, start=1):
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.fill = HEADER_FILL
        cell.border = Border()
        sheet.column_dimensions[get_column_letter(index)].width = COLUMNS[index - 1][2]

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return sheet


def existing_task_numbers(path: Path) -> set[str]:
    """أرقام المهام الموجودة أصلا في الملف، لمنع تكرار نفس المعاملة عند إعادة التشغيل."""
    if not path.exists():
        return set()

    workbook = load_workbook(path, read_only=True)
    seen: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2, min_col=TASK_COLUMN, max_col=TASK_COLUMN):
                value = row[0].value
                if value is not None and str(value).strip():
                    seen.add(str(value).strip())
    finally:
        workbook.close()
    return seen


def append_rows(path: Path, day_key: str, records: list[dict]) -> int:
    if not records:
        return 0

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _open_workbook(path)
    sheet = _get_day_sheet(workbook, day_key)

    serial = max(sheet.max_row - 1, 0)

    for record in records:
        serial += 1
        values = {**record, "serial": serial}
        sheet.append([values.get(key, "") for key in KEYS])

        row = sheet[sheet.max_row]
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)

        # تمييز الصفوف الناقصة بلون، حتى تعرف بنظرة واحدة ما يحتاج مراجعة يدوية.
        if not record.get("patientName") or not record.get("fileNumber"):
            for cell in row:
                cell.fill = NEEDS_REVIEW_FILL

    workbook.save(path)
    return len(records)


def update_statuses(path: Path, day_key: str, status_by_task: dict[str, str]) -> int:
    """
    يحدث خانة «حالة المعاملة» بعد تنفيذ خطوة "تحت التنفيذ" على النظام،
    فيبقى الإكسل معبرا عن الحالة الفعلية لا عن النية فقط.
    """
    if not path.exists() or not status_by_task:
        return 0

    workbook = load_workbook(path)
    if day_key not in workbook.sheetnames:
        workbook.close()
        return 0

    sheet = workbook[day_key]
    updated = 0

    for row_index in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=TASK_COLUMN).value
        task_number = str(value).strip() if value is not None else ""
        if task_number and task_number in status_by_task:
            sheet.cell(row=row_index, column=STATUS_COLUMN).value = status_by_task[task_number]
            updated += 1

    if updated:
        workbook.save(path)
    workbook.close()
    return updated
